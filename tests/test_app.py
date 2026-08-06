import base64
import os
import sqlite3
import tempfile
import unittest
from datetime import datetime, timedelta, timezone
from pathlib import Path
from io import BytesIO
from unittest import mock

import requests
from openpyxl import load_workbook

from issue_tracker.application import (
    create_app,
    ensure_issue_columns,
    get_connection,
    github_request,
    load_env_file,
    next_link_url,
)


class IssueTrackerTestCase(unittest.TestCase):
    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.app = create_app(
            {
                "TESTING": True,
                "DB_PATH": str(Path(self.temp_dir.name) / "test.db"),
                "APP_USERNAME": "tester",
                "APP_PASSWORD": "secret",
                "GITHUB_SSL_VERIFY": True,
            }
        )
        self.client = self.app.test_client()
        token = base64.b64encode(b"tester:secret").decode()
        self.headers = {"Authorization": f"Basic {token}"}
        now = datetime.now(timezone.utc).replace(microsecond=0)
        old = now - timedelta(days=60)
        with get_connection(self.app) as connection:
            connection.executemany(
                """
                INSERT INTO issues (
                    number, repository, title, body, html_url, upstream_state,
                    author, labels_json, github_created_at, github_updated_at,
                    comment_count, first_synced_at, last_synced_at,
                    identification_result, value_level, missed_test_reason,
                    supplemental_test, notes
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                [
                    (
                        101,
                        "vllm-project/vllm-ascend",
                        "Recent failure",
                        "reproduction",
                        "https://github.com/example/issues/101",
                        "open",
                        "alice",
                        '["bug"]',
                        now.isoformat(),
                        now.isoformat(),
                        1,
                        now.isoformat(),
                        now.isoformat(),
                        "确认问题",
                        "高",
                        "未覆盖并发组合",
                        "增加并发回归测试",
                        "优先分析",
                    ),
                    (
                        2,
                        "vllm-project/vllm-ascend",
                        "Old closed issue",
                        "old body",
                        "https://github.com/example/issues/2",
                        "closed",
                        "bob",
                        "[]",
                        old.isoformat(),
                        old.isoformat(),
                        0,
                        now.isoformat(),
                        now.isoformat(),
                        "",
                        "",
                        "",
                        "",
                        "",
                    ),
                ],
            )

    def tearDown(self):
        self.temp_dir.cleanup()

    def test_authentication_is_required(self):
        response = self.client.get("/api/issues")
        self.assertEqual(response.status_code, 401)

    def test_index_is_served_from_package_static_directory(self):
        response = self.client.get("/", headers=self.headers)
        self.assertEqual(response.status_code, 200)
        self.assertIn("vLLM Ascend Issue", response.get_data(as_text=True))
        self.assertIn("列设置", response.get_data(as_text=True))
        self.assertIn("问题分析", response.get_data(as_text=True))
        self.assertIn("漏测原因", response.get_data(as_text=True))
        self.assertIn("补充测试", response.get_data(as_text=True))
        self.assertNotIn('class="filter-band"', response.get_data(as_text=True))
        self.assertEqual(response.get_data(as_text=True).count('id="clearFilters"'), 1)

    def test_recent_identified_filter(self):
        response = self.client.get(
            "/api/issues?created=last_month&identified=确认问题",
            headers=self.headers,
        )
        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        self.assertEqual(payload["total"], 1)
        self.assertEqual(payload["items"][0]["number"], 101)

    def test_default_order_is_created_time_descending(self):
        response = self.client.get("/api/issues", headers=self.headers)
        numbers = [item["number"] for item in response.get_json()["items"]]
        self.assertEqual(numbers, [101, 2])

    def test_column_text_filters_match_full_dataset(self):
        response = self.client.get(
            "/api/issues?missed=并发&supplemental=回归&notes=优先",
            headers=self.headers,
        )
        payload = response.get_json()
        self.assertEqual(payload["total"], 1)
        self.assertEqual(payload["items"][0]["number"], 101)

    def test_excel_export_uses_filters_and_visible_columns(self):
        response = self.client.get(
            "/api/issues/export?missed=并发&columns=issue,missed,notes",
            headers=self.headers,
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.mimetype,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        workbook = load_workbook(BytesIO(response.data), read_only=True)
        rows = list(workbook.active.iter_rows(values_only=True))
        self.assertEqual(rows[0], ("Issue", "漏测原因", "备注"))
        self.assertEqual(len(rows), 2)
        self.assertIn("#101", rows[1][0])

    def test_excel_export_can_contain_only_issue_column(self):
        response = self.client.get(
            "/api/issues/export?state=closed&columns=issue",
            headers=self.headers,
        )
        workbook = load_workbook(BytesIO(response.data), read_only=True)
        rows = list(workbook.active.iter_rows(values_only=True))
        self.assertEqual(rows[0], ("Issue",))
        self.assertEqual(len(rows), 2)
        self.assertIn("#2", rows[1][0])

    def test_excel_export_includes_version_and_ai_analysis_columns(self):
        self.client.patch(
            "/api/issues/101",
            headers=self.headers,
            json={
                "affected_version": "v0.11.0rc1",
                "version_support_status": "下个版本支持",
                "ai_analysis": "## 结论\n\n需要补充回归测试",
            },
        )
        response = self.client.get(
            "/api/issues/export?columns=issue,version,version_support,ai_analysis",
            headers=self.headers,
        )
        workbook = load_workbook(BytesIO(response.data), read_only=True)
        rows = list(workbook.active.iter_rows(values_only=True))
        self.assertEqual(
            rows[0], ("Issue", "问题版本", "版本支持情况", "AI分析结论")
        )
        self.assertEqual(rows[1][1:], (
            "v0.11.0rc1",
            "下个版本支持",
            "## 结论\n\n需要补充回归测试",
        ))

    def test_manual_analysis_can_be_updated(self):
        response = self.client.patch(
            "/api/issues/101",
            headers=self.headers,
            json={
                "summary_zh": "可复现的启动失败",
                "missed_test_reason": "未覆盖对应配置组合",
                "is_closed_loop": "是",
                "affected_version": "v0.11.0rc1",
                "version_support_status": "下个版本支持",
                "ai_analysis": "## 结论\n\n- 可稳定复现",
                "title": "must not be changed",
            },
        )
        self.assertEqual(response.status_code, 200)
        detail = self.client.get("/api/issues/101", headers=self.headers).get_json()
        self.assertEqual(detail["summary_zh"], "可复现的启动失败")
        self.assertEqual(detail["is_closed_loop"], "是")
        self.assertEqual(detail["affected_version"], "v0.11.0rc1")
        self.assertEqual(detail["version_support_status"], "下个版本支持")
        self.assertIn("<h2>结论</h2>", detail["ai_analysis_html"])
        self.assertEqual(detail["title"], "Recent failure")

        filtered = self.client.get(
            "/api/issues?closed_loop=是", headers=self.headers
        ).get_json()
        self.assertEqual(filtered["total"], 1)

        filtered = self.client.get(
            "/api/issues?version=v0.11&version_support=下个版本支持&ai_analysis=稳定复现",
            headers=self.headers,
        ).get_json()
        self.assertEqual(filtered["total"], 1)

    def test_existing_database_is_migrated_without_losing_rows(self):
        connection = sqlite3.connect(":memory:")
        connection.row_factory = sqlite3.Row
        connection.execute("CREATE TABLE issues (number INTEGER PRIMARY KEY)")
        connection.execute("INSERT INTO issues(number) VALUES (7)")

        ensure_issue_columns(connection)

        row = connection.execute(
            """
            SELECT number, is_closed_loop, affected_version,
                   version_support_status, ai_analysis
            FROM issues WHERE number = 7
            """
        ).fetchone()
        self.assertEqual(row["number"], 7)
        self.assertEqual(row["is_closed_loop"], "")
        self.assertEqual(row["affected_version"], "")
        self.assertEqual(row["version_support_status"], "")
        self.assertEqual(row["ai_analysis"], "")
        connection.close()

    def test_invalid_version_support_status_is_rejected(self):
        response = self.client.patch(
            "/api/issues/101",
            headers=self.headers,
            json={"version_support_status": "随便填写"},
        )
        self.assertEqual(response.status_code, 400)
        self.assertIn("版本支持情况", response.get_json()["error"])

    def test_markdown_preview_is_rendered_and_sanitized(self):
        response = self.client.post(
            "/api/markdown/render",
            headers=self.headers,
            json={
                "markdown": "# 分析\n\n|场景|结果|\n|-|-|\n|并发|失败|\n\n"
                "[危险链接](javascript:alert(1))<script>alert(2)</script>"
            },
        )
        self.assertEqual(response.status_code, 200)
        rendered = response.get_json()["html"]
        self.assertIn("<h1>分析</h1>", rendered)
        self.assertIn("<table>", rendered)
        self.assertNotIn("javascript:", rendered)
        self.assertNotIn("<script", rendered)

    def test_cursor_pagination_uses_next_link(self):
        link = (
            '<https://api.github.com/repos/example/repo/issues?after=abc>; rel="next", '
            '<https://api.github.com/repos/example/repo/issues?before=xyz>; rel="prev"'
        )
        self.assertEqual(
            next_link_url(link),
            "https://api.github.com/repos/example/repo/issues?after=abc",
        )

    def test_github_ssl_verification_can_be_disabled(self):
        self.app.config["GITHUB_SSL_VERIFY"] = False
        response = mock.MagicMock()
        response.status_code = 200
        response.json.return_value = []
        response.headers = {}

        with mock.patch.object(
            self.app.extensions["github_http"],
            "get",
            return_value=response,
        ) as get:
            github_request(self.app)

        self.assertFalse(get.call_args[1]["verify"])

    def test_incomplete_github_response_is_retried(self):
        response = mock.MagicMock()
        response.status_code = 200
        response.json.return_value = []
        response.headers = {}

        with mock.patch.object(
            self.app.extensions["github_http"],
            "get",
            side_effect=[
                requests.exceptions.ChunkedEncodingError("partial response"),
                response,
            ],
        ) as get, mock.patch("issue_tracker.application.time.sleep") as sleep:
            payload, _, _ = github_request(self.app)

        self.assertEqual(payload, [])
        self.assertEqual(get.call_count, 2)
        sleep.assert_called_once_with(1)

    def test_rate_limit_error_recommends_authenticated_token(self):
        response = mock.MagicMock()
        response.status_code = 403
        response.text = '{"message":"rate limit exceeded"}'
        response.headers = {
            "X-RateLimit-Remaining": "0",
            "X-RateLimit-Reset": "1234567890",
        }

        with mock.patch.object(
            self.app.extensions["github_http"],
            "get",
            return_value=response,
        ), self.assertRaisesRegex(RuntimeError, "GITHUB_TOKEN"):
            github_request(self.app)

    def test_env_file_is_loaded_without_overriding_process_environment(self):
        env_file = Path(self.temp_dir.name) / ".env"
        env_file.write_text(
            "APP_USERNAME=file-user\nGITHUB_SSL_VERIFY=false\n",
            encoding="utf-8",
        )

        with mock.patch.dict(os.environ, {"APP_USERNAME": "process-user"}, clear=True):
            load_env_file(env_file)
            self.assertEqual(os.environ["APP_USERNAME"], "process-user")
            self.assertEqual(os.environ["GITHUB_SSL_VERIFY"], "false")


if __name__ == "__main__":
    unittest.main()
